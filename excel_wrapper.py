from __future__ import annotations

import contextlib
import re
from collections import OrderedDict
from datetime import datetime
from pathlib import Path
from typing import Union

try:
    # Need openpyxl >= v2.6
    from openpyxl import Workbook
    from openpyxl.formatting.rule import Rule
    from openpyxl.styles import PatternFill
    from openpyxl.styles.differential import DifferentialStyle
    from openpyxl.worksheet.table import Table, TableStyleInfo
except ImportError as e:
    raise SystemExit('Requires "openpyxl" Package') from e

""" Example Usage:

    # {} automatically inserts YYYYMMDD in the filename
    excel = ExcelWrapper('~/my_aws_inventory-{}.xlsx')

    excel.add_headers('Instances', ['Region', 'Instance ID'])
    excel.add_headers('Volumes', ['Region', 'Volume ID'])

    excel.add_data('Instances', ['us-east-1', 'i-xxxxx1'])
    excel.add_data('Instances', ['eu-west-1', 'i-xxxxx2'])

    excel.add_data('Volumes', ['eu-west-1', 'vol-xxxxx1'])
    excel.add_data('Volumes', ['eu-west-1', 'vol-xxxxx2'])

    excel.sort_data('Instances', ['Region', 'Instance ID'])
    excel.sort_data('Volumes', 'Region')

    excel.export_excel()
"""


def natural_sort_multiple_columns(data: list[list], sort_keys: list) -> None:
    """In-place natural sorting of a nested list

    Args:
        data (list[list]): Nested List
        sort_keys (list): Sort Indexes
    """

    def natural_sort_key(s):
        return [
            int(text) if text.isdigit() else text.lower()
            for text in re.split('([0-9]+)', s)
        ]

    data.sort(key=lambda x: [natural_sort_key(str(x[key])) for key in sort_keys])


class ExcelWrapper(object):
    output_file = None
    output_data = OrderedDict()
    table_styles = {}
    cell_styles = {}
    frozen_columns = {}

    def __init__(self, output_file: str, table_style: str = 'TableStyleMedium9'):
        self.table_style = table_style
        self.output_file = Path(
            output_file.format(datetime.now().strftime('%Y%m%d'))
        ).expanduser()
        self.output_data = {}

    @staticmethod
    def _col_to_excel(col):
        excel_col = str()
        div = col
        while div:
            (div, mod) = divmod(div - 1, 26)
            excel_col = chr(mod + 65) + excel_col

        return excel_col

    def _autosize_columns(self, current_sheet):
        # Autosize the columns
        column_adjustment = 1.12
        for col in current_sheet.columns:
            max_length = 0
            column = col[0].column_letter  # Get the column name
            for cell in col:
                # If this is a cell w/ a date, lets format it
                if cell.data_type == 'd':
                    # Strip the timezone
                    cell.value = cell.value.replace(tzinfo=None)

                    cell.number_format = 'yyyy-mm-dd hh:mm:ss'
                    max_length = 16

                with contextlib.suppress(Exception):
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)

            adjusted_width = (max_length + 2) * column_adjustment
            current_sheet.column_dimensions[column].width = adjusted_width

    def _create_sheet(self, sheet_name: str):
        if sheet_name not in self.output_data:
            self.output_data[sheet_name] = {'headers': [], 'data': []}

    def _freeze_columns(self, current_sheet):
        sheet_name = current_sheet.title

        if sheet_name in self.frozen_columns:
            sheet_headers = next(current_sheet.rows)
            try:
                column_number = next(
                    i.col_idx
                    for i in sheet_headers
                    if i.value == self.frozen_columns[sheet_name]
                )
            except StopIteration as e:
                raise RuntimeError('Invalid Header') from e
            else:
                cell = f'{self._col_to_excel(column_number + 1)}1'
                current_sheet.freeze_panes = current_sheet[cell]

    def _merge_data(self, sheet_name: str) -> list:
        return_data = [self.output_data.get(sheet_name).get('headers')]

        # Validate header count
        header_count = len(self.output_data.get(sheet_name).get('headers'))

        if not header_count:
            raise SystemExit(f'No Headers Found: {sheet_name}')

        for data in self.output_data.get(sheet_name).get('data'):
            if header_count != len(data):
                raise SystemExit('Invalid Data Found, Column Count Mismatch')

            return_data.append(data)

        return return_data

    def add_headers(self, sheet_name: str, data: Union[list, tuple, set]) -> None:
        self._create_sheet(sheet_name)

        data_list = list(data)  # Convert to list
        if len(data_list) != len(set(data_list)):
            raise ValueError('Duplicate Header Found.')
        self.output_data[sheet_name]['headers'] = data_list

    def add_data(self, sheet_name: str, data: Union[list, tuple, set]) -> None:
        self._create_sheet(sheet_name)

        if not isinstance(data, (list, tuple, set)):
            raise TypeError('Invalid Data Type, Requires List, Tuple, or Set')

        if isinstance(data[0], (list, tuple, set)):
            # It's batch data
            for batch_row in data:
                if len(batch_row) != len(self.output_data[sheet_name]['headers']):
                    raise ValueError('Invalid Data Found, Column Count Mismatch')
                self.output_data[sheet_name]['data'].append(
                    [item if item is not None else "" for item in batch_row]
                )
        elif len(data) == len(self.output_data[sheet_name]['headers']):
            self.output_data[sheet_name]['data'].append(
                [item if item is not None else "" for item in data]
            )
        else:
            raise ValueError('Invalid Data Found, Column Count Mismatch')

    def sort_data(self, sheet_name: str, headers: Union[list, tuple, set]) -> None:
        sort_keys = []

        if sheet_name not in self.output_data:
            raise SystemExit('Cannot sort where no headers have been found.')

        headers = [headers] if isinstance(headers, str) else list(headers)

        for header in headers:
            sheet_headers = self.output_data.get(sheet_name).get('headers')
            if header not in sheet_headers:
                raise SystemExit(f'Invalid Sort Header: {header}')

            sort_keys.append(sheet_headers.index(header))

        self.output_data[sheet_name]['sort_keys'] = sort_keys

    def format_cells(
        self, sheet_name: str, header: str, rule: str, bg_color: str
    ) -> None:
        self.cell_styles.setdefault(sheet_name, {})
        self.cell_styles[sheet_name].setdefault(header, []).append(
            {'rule': rule, 'bg_color': bg_color}
        )

    def conditional_format(self, sheet_name: str, rules: list[dict]) -> None:
        """Add Conditional Formatting to Table

        Args:
            sheet_name (str): name of sheet
            rules (list[dict]): list of rules

        Required fields inside nested dictionaries:
            - formula
            - bg_color
        """
        if not (
            isinstance(rules, list) and all(isinstance(value, dict) for value in rules)
        ):
            print('List of Dicts Reqd for Conditional Formatting - Ignoring!')
            return

        self.table_styles.setdefault(sheet_name, [])

        for rule in rules:
            formula = rule.get('formula')
            bg_color = rule.get('bg_color')

            if not (formula and bg_color):
                print('Skipping Formatting - Missing formula and/or bg_color!')
                continue

            self.table_styles[sheet_name].append(
                Rule(
                    type="expression",
                    dxf=DifferentialStyle(
                        fill=PatternFill(bgColor=bg_color.lstrip('#'))
                    ),
                    formula=formula,
                    stopIfTrue=False,
                )
            )

    def freeze_column_after(self, sheet_name: str, column_name: str) -> None:
        self.frozen_columns[sheet_name] = column_name

    def export_excel(self):
        # Setup the output XLSX and remove default sheet
        excel_workbook = Workbook()
        excel_workbook.remove(excel_workbook.active)

        for sheet_name, sheet_data in self.output_data.items():
            if not sheet_data.get('data'):
                print(f'{sheet_name}: No Data Found')
                continue

            current_sheet = excel_workbook.create_sheet(sheet_name)

            if sort_keys := sheet_data.get('sort_keys'):
                natural_sort_multiple_columns(
                    self.output_data[sheet_name]['data'], sort_keys
                )

            for item in self._merge_data(sheet_name):
                current_sheet.append(item)

            max_column = self._col_to_excel(current_sheet.max_column)
            max_row = current_sheet.max_row

            current_sheet.add_table(
                Table(
                    displayName=sheet_name.replace(' ', ''),
                    ref=f'A1:{max_column}{max_row}',
                    tableStyleInfo=TableStyleInfo(
                        name=self.table_style,
                        showFirstColumn=False,
                        showLastColumn=False,
                        showRowStripes=True,
                        showColumnStripes=False,
                    ),
                )
            )

            for header, cell_rules in self.cell_styles.get(sheet_name, {}).items():
                for rule in cell_rules:
                    header_idx = (
                        self.output_data.get(sheet_name).get('headers').index(header)
                        + 1
                    )
                    header_col = self._col_to_excel(header_idx)

                    header_ref = f'{header_col}2:{header_col}{max_row}'
                    formula = rule['rule']
                    bg_color = rule['bg_color']

                    rule = Rule(
                        type="expression",
                        dxf=DifferentialStyle(
                            fill=PatternFill(bgColor=bg_color.lstrip('#'))
                        ),
                        formula=[formula],
                        stopIfTrue=False,
                    )

                    current_sheet.conditional_formatting.add(header_ref, rule)

            # Add conditional formatting to sheet
            for style in self.table_styles.get(sheet_name, []):
                current_sheet.conditional_formatting.add(
                    f'A2:{max_column}{max_row}', style
                )

            self._autosize_columns(current_sheet)
            self._freeze_columns(current_sheet)

        try:
            excel_workbook.save(self.output_file)
        except IndexError:
            print('FATAL: No Sheets Containing Data - No Excel Doc Generated!')
            return
        finally:
            excel_workbook.close()


if __name__ == "__main__":
    excel = ExcelWrapper('~/my_aws_inventory-{}.xlsx')

    excel.add_headers('Instances', ['Region', 'Instance ID'])
    excel.add_headers('Volumes', ['Region', 'Volume ID'])
    excel.add_headers('RDS', ['Region', 'Volume ID'])

    excel.add_data('Instances', ['us-east-1', 'i-xxxxx1'])
    excel.add_data('Instances', ['eu-west-1', 'i-xxxxx2'])

    excel.add_data(
        'Volumes',
        [
            ['eu-west-1', 'vol-xxxxx1'],
            ['eu-west-1', 'vol-xxxxx2'],
        ],
    )

    excel.add_data('RDS', ['us-east-1', 'i-xxxxx1'])
    excel.add_data('RDS', ['eu-west-1', 'i-xxxxx2'])

    excel.format_cells(
        sheet_name='Instances',
        header='Region',
        rule='$A2="eu-west-1"',
        bg_color='#E6B8B7',
    )

    excel.conditional_format(
        'Volumes',
        [
            {
                'formula': ['$A2 = "eu-west-1"'],
                'bg_color': '#E6B8B7',
                'row_condition': True,
            }
        ],
    )

    excel.sort_data('Instances', ('Region'))
    excel.export_excel()
